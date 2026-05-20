import os
import tempfile
from sqlalchemy.ext.asyncio import AsyncSession
from repositories.file_repo import FileRepo
from config import settings


class DocumentService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.file_repo = FileRepo(db)

    async def parse(self, file_id: str, options: dict | None = None) -> dict:
        doc = await self.file_repo.get(file_id)
        if not doc:
            return None
        full_path = os.path.join(settings.workspace_dir, doc.file_path)
        if not os.path.isfile(full_path):
            return {"file_id": file_id, "content": doc.content_text or "", "tables": [], "metadata": {}}

        ext = doc.file_type.lower()
        content = ""
        tables = []

        if ext == "pdf":
            from pypdf import PdfReader
            reader = PdfReader(full_path)
            pages = []
            for page in reader.pages:
                pages.append(page.extract_text() or "")
            content = "\n".join(pages)
        elif ext == "docx":
            from docx import Document as DocxDoc
            d = DocxDoc(full_path)
            content = "\n".join(p.text for p in d.paragraphs)
            if options and options.get("extract_tables"):
                for table in d.tables:
                    rows = [[cell.text for cell in row.cells] for row in table.rows]
                    tables.append({"header": rows[0] if rows else [], "rows": rows[1:] if len(rows) > 1 else []})
        elif ext in ("xls", "xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(full_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
                if rows:
                    tables.append({"header": rows[0], "rows": rows[1:]})
            content = "\n".join(" | ".join(row) for table in tables for row in [table["header"]] + table["rows"])
        elif ext in ("txt", "md", "markdown"):
            with open(full_path, "r", encoding="utf-8") as f:
                content = f.read()
        elif ext in ("htm", "html"):
            from bs4 import BeautifulSoup
            with open(full_path, "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f.read(), "html.parser")
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                content = soup.get_text(separator="\n")
        else:
            content = doc.content_text or ""

        doc.content_text = content
        await self.db.commit()
        return {
            "file_id": file_id,
            "content": content[:50000],
            "tables": tables,
            "metadata": {"pages": content.count("\f") + 1, "word_count": len(content.split())},
        }

    async def parse_url(self, url: str, extract_images: bool = False) -> dict:
        import httpx
        from bs4 import BeautifulSoup
        async with httpx.AsyncClient() as client:
            resp = await client.get(url, timeout=15, follow_redirects=True)
            soup = BeautifulSoup(resp.text, "html.parser")
            title = soup.title.string if soup.title else url
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            content = soup.get_text(separator="\n")
            return {
                "url": url,
                "title": title,
                "content": content[:50000],
                "metadata": {"author": "", "published_at": ""},
            }
